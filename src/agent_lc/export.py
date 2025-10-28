# src/agent_lc/export.py
import pandas as pd, sqlite3, openpyxl
from openpyxl.styles import Alignment, PatternFill
from .config import DB_PATH
import numpy as np

def _build_by_id(df: pd.DataFrame):
    # устойчивый маппинг id -> строка
    df = df.copy().reset_index(drop=True)
    n = len(df)
    if "id" not in df.columns:
        df["id"] = pd.Series(range(1, n + 1), index=df.index, dtype=int)
    else:
        ids = pd.to_numeric(df["id"], errors="coerce")
        fallback = pd.Series(range(1, n + 1), index=df.index, dtype=float)
        ids = ids.fillna(fallback)
        # защитимся от нулей/отрицательных/inf
        bad = ~np.isfinite(ids) | (ids <= 0)
        if bad.any():
            repl = pd.Series(range(1, n + 1), index=df.index, dtype=float)
            ids[bad] = repl[bad]
        df["id"] = ids.astype(int)
    return {int(r["id"]): r for _, r in df.iterrows()}

def _fmt_bool(x):
    return int(bool(x)) if pd.notna(x) else 0

def export_excel_report(df_scored: pd.DataFrame, llm_resp: dict, file_path: str) -> str:
    by_id = _build_by_id(df_scored)

    # основной лист risk
    rows = []
    for t in llm_resp.get("transactions", []):
        rid = t.get("id")
        base = by_id.get(int(rid)) if rid is not None else {}
        ev = t.get("evidence", {}) or {}

        rows.append({
            "id": rid,
            "date": base.get("date") or base.get("ts"),
            "debit_account": base.get("debit_account"),
            "debit_name": base.get("debit_name"),
            "debit_inn": base.get("debit_inn"),
            "credit_account": base.get("credit_account"),
            "credit_name": base.get("credit_name"),
            "credit_inn": base.get("credit_inn"),

            "purpose": t.get("purpose", ""),

            # итог и компоненты
            "risk_label": t.get("risk_label", ""),
            "risk_score": float(t.get("risk_score", 0.0) or 0.0),
            "p_ml": ev.get("ml_metric", base.get("ml_metric")),
            "p_prior": ev.get("prior", None),
            "p_llm": ev.get("p_llm", None),
            "p_final": ev.get("p_final", t.get("risk_score")),

            # правила/флаги/объяснения
            "rule_hits": ", ".join(t.get("rule_hits", [])) if isinstance(t.get("rule_hits"), list) else (t.get("rule_hits") or ""),
            "flags": ", ".join(t.get("flags", []) or []),
            "primary_reasons": "; ".join(t.get("primary_reasons", []) or []),
            "recommendation": t.get("recommendation", ""),
            "risk_explanation": t.get("risk_explanation", ""),

            # память по контрагентам (сводно)
            "debit_susp_rate": base.get("debit_susp_rate"),
            "debit_cnt_suspicious": base.get("debit_cnt_suspicious"),
            "debit_last_seen_days": base.get("debit_last_seen_days"),
            "debit_watchlisted": _fmt_bool(base.get("debit_watchlisted")),
            "debit_p95": base.get("debit_p95"),

            "credit_susp_rate": base.get("credit_susp_rate"),
            "credit_cnt_suspicious": base.get("credit_cnt_suspicious"),
            "credit_last_seen_days": base.get("credit_last_seen_days"),
            "credit_watchlisted": _fmt_bool(base.get("credit_watchlisted")),
            "credit_p95": base.get("credit_p95"),
        })

    out = pd.DataFrame(rows)

    with pd.ExcelWriter(file_path, engine="openpyxl") as wr:
        # ----- Лист risk
        out.to_excel(wr, index=False, sheet_name="risk")
        ws = wr.sheets.get("risk") or wr.book["risk"]

        # wrap для длинных текстов
        wrap_cols = [c for c in ["recommendation","risk_explanation","purpose","primary_reasons","flags"] if c in out.columns]
        for col_name in wrap_cols:
            col_idx = out.columns.get_loc(col_name) + 1
            for col_cells in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, max_row=ws.max_row):
                for cell in col_cells:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

        # подсветка по risk_label
        color = {"красный":"FFC7CE","желтый":"FFEB9C","жёлтый":"FFEB9C","зеленый":"C6EFCE","зелёный":"C6EFCE"}
        if "risk_label" in out.columns:
            rcol = out.columns.get_loc("risk_label") + 1
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                label = row[rcol-1].value
                if label in color:
                    fill = PatternFill(start_color=color[label], end_color=color[label], fill_type="solid")
                    for cell in row: cell.fill = fill

        # авто-ширина
        for i, col_name in enumerate(out.columns, start=1):
            max_len = max([len(str(col_name))] + [len(str(c.value)) if c.value is not None else 0
                                                  for c in ws.iter_cols(min_col=i, max_col=i, min_row=1, max_row=ws.max_row).__next__()])
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = min(60, max_len + 2)

        # ----- Лист review_queue (LLM «красный», но итог < 0.40)
# ----- Лист review_queue
        if not out.empty:
            def _is_llm_red(row):
                p_llm = row.get("p_llm")
                try:
                    return float(p_llm) >= 0.99
                except:
                    return False

            # p_final может отсутствовать — подстрахуемся risk_score
            p_final_series = out.get("p_final")
            if p_final_series is None:
                p_final_series = out["risk_score"]
            else:
                p_final_series = p_final_series.fillna(out["risk_score"])

            # 1) Разногласие: LLM == красный, система == зелёный
            disagree_mask = (out.apply(_is_llm_red, axis=1)) & (p_final_series < 0.40)

            # 2) Жёлтые случаи: 0.40 ≤ p_final < 0.70 (или по label)
            yellow_mask = out["risk_label"].isin(["желтый", "жёлтый"])
            # (если хочешь по числу, то так: yellow_mask = (p_final_series >= 0.40) & (p_final_series < 0.70))

            # Итоговая корзина на ручной обзор
            review_mask = disagree_mask | yellow_mask
            review = out.loc[review_mask].copy()

            review.to_excel(wr, index=False, sheet_name="review_queue")

            # немного wrap и авто-ширина
            ws2 = wr.sheets.get("review_queue") or wr.book["review_queue"]
            wrap_cols = [c for c in ["recommendation","risk_explanation","purpose","primary_reasons","flags"] if c in review.columns]
            for col_name in wrap_cols:
                col_idx = review.columns.get_loc(col_name) + 1
                for col_cells in ws2.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, max_row=ws2.max_row):
                    for cell in col_cells:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")

            for i, col_name in enumerate(review.columns, start=1):
                max_len = max([len(str(col_name))] + [
                    len(str(c.value)) if c.value is not None else 0
                    for c in ws2.iter_cols(min_col=i, max_col=i, min_row=1, max_row=ws2.max_row).__next__()
                ])
                ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = min(60, max_len + 2)

        # ----- Лист memory_summary (топ по памяти, с мягкими LLM-флагами)
        try:
            con = sqlite3.connect(DB_PATH); cur = con.cursor()
            top_agg = cur.execute("""
                SELECT inn, cnt_total, cnt_suspicious,
                       ROUND(CASE WHEN cnt_total>0 THEN 100.0*cnt_suspicious/cnt_total ELSE 0 END, 1) AS susp_rate_pct,
                       amt_total, amt_suspicious, last_seen_ts, watchlisted, p50, p75, p90, p95,
                       llm_flags_total, llm_last_seen_ts
                FROM agg_counterparty
                ORDER BY cnt_suspicious DESC, susp_rate_pct DESC
                LIMIT 200
            """).fetchall()
            con.close()
            cols = ["inn","cnt_total","cnt_suspicious","susp_rate_pct","amt_total","amt_suspicious",
                    "last_seen_ts","watchlisted","p50","p75","p90","p95","llm_flags_total","llm_last_seen_ts"]
            pd.DataFrame(top_agg, columns=cols).to_excel(wr, index=False, sheet_name="memory_summary")
        except Exception:
            pd.DataFrame(columns=["inn","cnt_total","cnt_suspicious","susp_rate_pct",
                                  "amt_total","amt_suspicious","last_seen_ts","watchlisted",
                                  "p50","p75","p90","p95","llm_flags_total","llm_last_seen_ts"]
                         ).to_excel(wr, index=False, sheet_name="memory_summary")

    return file_path
